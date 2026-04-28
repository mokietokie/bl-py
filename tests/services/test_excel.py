import tempfile
from pathlib import Path
import pytest
from openpyxl import Workbook, load_workbook
from bl_tracker.db import connection, repo
from bl_tracker.services import excel


@pytest.fixture
def db():
    with tempfile.TemporaryDirectory() as d:
        p = Path(d) / "t.sqlite"
        connection.init(p)
        yield p


def make_xlsx(tmp_path, rows):
    wb = Workbook()
    ws = wb.active
    ws.append(["BL번호", "IMO번호", "ETA", "화물위치"])
    for r in rows:
        ws.append(r)
    out = tmp_path / "in.xlsx"
    wb.save(out)
    return out


def test_import_creates_rows(tmp_path, db):
    f = make_xlsx(tmp_path, [["BL1", "1234567", "ETA-A", "Off Busan"],
                             ["BL2", "7654321", "ETA-B", "ECS"]])
    n = excel.import_xlsx(db, f)
    assert n == 2
    assert len(repo.list_shipments(db)) == 2


def test_import_upserts_on_bl(tmp_path, db):
    repo.create_shipment(db, bl_no="BL1", imo_no="OLD")
    f = make_xlsx(tmp_path, [["BL1", "NEW", "", ""]])
    excel.import_xlsx(db, f)
    s = [r for r in repo.list_shipments(db) if r["bl_no"] == "BL1"][0]
    assert s["imo_no"] == "NEW"
    assert len(repo.list_shipments(db)) == 1


def test_export_roundtrip(tmp_path, db):
    repo.create_shipment(db, bl_no="BL1", imo_no="111")
    out = tmp_path / "out.xlsx"
    excel.export_xlsx(db, out)
    wb = load_workbook(out)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers[:6] == ["BL번호", "선사", "선박명", "IMO번호", "ETA", "화물위치"]
    assert ws.cell(row=2, column=1).value == "BL1"


def test_export_roundtrip_through_import(tmp_path, db):
    repo.create_shipment(db, bl_no="BL1", imo_no="111", memo="hi")
    out = tmp_path / "out.xlsx"
    excel.export_xlsx(db, out)
    # wipe DB and re-import the exported file
    repo.delete_shipment(db, repo.list_shipments(db)[0]["id"])
    n = excel.import_xlsx(db, out)
    assert n == 1
    s = repo.list_shipments(db)[0]
    assert s["bl_no"] == "BL1"
    assert s["imo_no"] == "111"
    assert s["memo"] == "hi"
