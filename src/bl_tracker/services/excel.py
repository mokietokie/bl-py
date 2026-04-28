from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook, load_workbook

from bl_tracker.db import repo

HEADERS = [
    "BL번호", "선사", "선박명", "IMO번호", "ETA", "화물위치",
    "이전 ETA", "변경", "갱신시각(BL)", "갱신시각(위치)", "메모",
]


def _idx_of(header_row: tuple, *names: str) -> int | None:
    for i, cell in enumerate(header_row):
        if cell in names:
            return i
    return None


def import_xlsx(db: Path, file: Path) -> int:
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0
    header = rows[0]
    bl_idx = _idx_of(header, "BL번호") or 0
    imo_idx = _idx_of(header, "IMO번호", "IMO")
    if imo_idx is None:
        imo_idx = 1
    memo_idx = _idx_of(header, "메모")
    count = 0
    for r in rows[1:]:
        if not r or r[bl_idx] in (None, ""):
            continue
        bl_no = str(r[bl_idx]).strip()
        imo_no = (str(r[imo_idx]).strip()
                  if imo_idx is not None and len(r) > imo_idx and r[imo_idx] not in (None, "")
                  else None)
        memo = (str(r[memo_idx]).strip()
                if memo_idx is not None and len(r) > memo_idx and r[memo_idx] not in (None, "")
                else None)
        repo.upsert_shipment_by_bl(db, bl_no=bl_no, imo_no=imo_no, memo=memo)
        count += 1
    return count


def export_xlsx(db: Path, file: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "shipments"
    ws.append(HEADERS)
    for s in repo.list_shipments(db):
        ws.append([
            s["bl_no"], s["carrier"], s["vessel"], s["imo_no"],
            s["eta"], s["location"],
            s["eta_prev_kst"], "Y" if s["eta_changed"] else "",
            s["bl_refreshed_at"], s["loc_refreshed_at"], s["memo"],
        ])
    wb.save(file)
    return file
